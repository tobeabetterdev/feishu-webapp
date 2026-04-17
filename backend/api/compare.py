from __future__ import annotations

import asyncio
import base64
import io
import json
import logging
import math
import os
import posixpath
import re
import traceback
import uuid
import warnings
import zipfile
from xml.etree import ElementTree as ET
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import quote

import pandas as pd
from openpyxl import load_workbook
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

from services.hengyi_order_comparison import (
    HengyiComparisonError,
    compare_hengyi_data,
    parse_hengyi_factory_data,
    parse_hengyi_jiuding_data,
)
from services.xinfengming_order_comparison import compare_xinfengming_data

router = APIRouter()
tasks: Dict[str, Dict[str, Any]] = {}
FACTORY_GROUPS_CONFIG_PATH = Path(__file__).resolve().parent.parent / "config" / "factory_groups.json"
RUNTIME_DIR = Path(__file__).resolve().parent.parent / "runtime"
RUNTIME_DIR.mkdir(exist_ok=True)
LOGGER = logging.getLogger("compare_tasks")
if not LOGGER.handlers:
    file_handler = logging.FileHandler(RUNTIME_DIR / "backend.tasks.log", encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    LOGGER.addHandler(file_handler)
    LOGGER.addHandler(stream_handler)
    LOGGER.setLevel(logging.INFO)

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)

HENGYI_DETAIL_COLUMNS = [
    "异常类型",
    "交货单",
    "送达方",
    "车牌号",
    "物料组",
    "工厂",
    "交货数量",
    "托盘数",
    "业务员",
    "过账日期",
    "出库单号",
    "会员名称",
    "产品类型",
    "客户名称",
    "子公司名称",
    "出库数量",
    "订单日期2",
    "出库数量差异",
]

HENGYI_HEADER_GROUPS = {
    "工厂": ["交货单", "送达方", "车牌号", "物料组", "工厂", "交货数量", "托盘数", "业务员", "过账日期"],
    "久鼎": ["出库单号", "会员名称", "产品类型", "客户名称", "子公司名称", "出库数量", "订单日期2"],
}

XINFENGMING_SUMMARY_COLUMNS = [
    "日期",
    "单号",
    "工厂",
    "型号",
    "公司",
    "客户出库数",
    "久鼎出库数",
    "待处理数量",
]

XINFENGMING_DETAIL_COLUMNS = [
    "异常类型",
    "交货单号",
    "交货创建日期",
    "销售组织描述",
    "客户名称",
    "业务员",
    "车牌号",
    "物料组描述",
    "件数",
    "交货单类型",
    "包装批号",
    "出库单号",
    "订单日期",
    "客户名称2",
    "会员名称",
    "产品类型",
    "实际出库数量",
    "子公司名称",
    "订单状态",
    "送货方式",
    "差异数量",
]

XINFENGMING_HEADER_GROUPS = {
    "工厂": ["交货单号", "交货创建日期", "销售组织描述", "客户名称", "业务员", "车牌号", "物料组描述", "件数", "交货单类型", "包装批号"],
    "久鼎": ["出库单号", "订单日期", "客户名称2", "会员名称", "产品类型", "实际出库数量", "子公司名称", "订单状态", "送货方式", "差异数量"],
}


class TaskStatus(BaseModel):
    task_id: str
    status: str
    progress: int
    message: str


def _ensure_excel(filename: str, label: str) -> None:
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail=f"{label}必须是 Excel 文件")


def _clean_data(data):
    if isinstance(data, dict):
        return {key: _clean_data(value) for key, value in data.items()}
    if isinstance(data, list):
        return [_clean_data(item) for item in data]
    if isinstance(data, float) and not math.isfinite(data):
        return None
    return data


_CELL_REF_PATTERN = re.compile(r"([A-Z]+)(\d+)")


def _column_ref_to_index(column_ref: str) -> int:
    index = 0
    for char in column_ref:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _read_xlsx_rows_from_zip(file_bytes: bytes) -> list[list[str | None]]:
    namespace = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        first_sheet = workbook_root.find("main:sheets/main:sheet", namespace)
        if first_sheet is None:
            return []

        relation_id = first_sheet.attrib.get(f"{{{namespace['rel']}}}id")
        if not relation_id:
            return []

        workbook_rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        target_path = None
        for relation in workbook_rels_root.findall("pkgrel:Relationship", namespace):
            if relation.attrib.get("Id") == relation_id:
                target_path = relation.attrib.get("Target")
                break

        if not target_path:
            return []

        normalized_target_path = target_path.lstrip("/")
        if not normalized_target_path.startswith("xl/"):
            normalized_target_path = posixpath.join("xl", normalized_target_path)
        target_path = posixpath.normpath(normalized_target_path)

        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("main:si", namespace):
                text_nodes = item.findall(".//main:t", namespace)
                shared_strings.append("".join(node.text or "" for node in text_nodes))

        sheet_root = ET.fromstring(archive.read(target_path))
        rows: list[list[str | None]] = []

        for row in sheet_root.findall("main:sheetData/main:row", namespace):
            row_values: list[str | None] = []
            for cell in row.findall("main:c", namespace):
                cell_ref = cell.attrib.get("r", "")
                match = _CELL_REF_PATTERN.match(cell_ref)
                if not match:
                    continue

                column_index = _column_ref_to_index(match.group(1))
                while len(row_values) <= column_index:
                    row_values.append(None)

                cell_type = cell.attrib.get("t")
                if cell_type == "inlineStr":
                    text_nodes = cell.findall(".//main:t", namespace)
                    row_values[column_index] = "".join(node.text or "" for node in text_nodes)
                    continue

                value_node = cell.find("main:v", namespace)
                if value_node is None:
                    row_values[column_index] = None
                    continue

                raw_value = value_node.text
                if cell_type == "s" and raw_value is not None:
                    shared_index = int(raw_value)
                    row_values[column_index] = shared_strings[shared_index] if shared_index < len(shared_strings) else raw_value
                else:
                    row_values[column_index] = raw_value

            rows.append(row_values)

    return rows


def _build_dataframe_from_rows(rows: list[list[str | None]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    normalized_rows = []
    width = len(header)
    for row in rows[1:]:
        current = list(row[:width])
        if len(current) < width:
            current.extend([None] * (width - len(current)))
        normalized_rows.append(current)

    return pd.DataFrame(normalized_rows, columns=header, dtype=str)


def _read_excel_with_fallback(file_bytes: bytes) -> pd.DataFrame:
    try:
        return pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except ValueError as exc:
        message = str(exc)
        if "Value must be either numerical or a string containing a wildcard" not in message:
            raise

        try:
            rows = _read_xlsx_rows_from_zip(file_bytes)
            return _build_dataframe_from_rows(rows)
        except zipfile.BadZipFile:
            raise exc


def _update_task(
    task_id: str,
    *,
    status: Optional[str] = None,
    progress: Optional[int] = None,
    message: Optional[str] = None,
) -> None:
    if status is not None:
        tasks[task_id]["status"] = status
    if progress is not None:
        tasks[task_id]["progress"] = progress
    if message is not None:
        tasks[task_id]["message"] = message

    LOGGER.info(
        "task=%s status=%s progress=%s message=%s",
        task_id,
        tasks[task_id].get("status"),
        tasks[task_id].get("progress"),
        tasks[task_id].get("message"),
    )


def _build_result_filename(result_df: pd.DataFrame, *, factory_type: str | None = None) -> str:
    date_for_name = ""
    for column_name in ("日期", "订单日期", "工厂过账日期", "久鼎订单日期"):
        if not result_df.empty and column_name in result_df.columns:
            first_date = result_df[column_name].dropna().iloc[0] if not result_df[column_name].dropna().empty else None
            if first_date is not None and str(first_date).strip():
                parsed_date = pd.to_datetime(first_date, errors="coerce")
                if pd.isna(parsed_date):
                    date_for_name = str(first_date).strip().replace("/", "-").split(" ")[0]
                else:
                    date_for_name = f"{parsed_date.year}-{parsed_date.month}-{parsed_date.day}"
                break

    factory_prefix_map = {
        "hengyi": "恒逸订单核对",
        "xinfengming": "新凤鸣订单核对",
    }
    filename_parts = [factory_prefix_map.get(factory_type or "", "订单核对")]
    if date_for_name:
        filename_parts.append(date_for_name)
    filename_parts.append(str(uuid.uuid4())[:8])
    return "_".join(filename_parts) + ".xlsx"


def _first_non_empty_value(series: pd.Series):
    for value in series:
        if value is None or pd.isna(value):
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def _build_hengyi_order_key(row: pd.Series) -> str:
    for column_name in ("订单号", "交货单", "出库单号"):
        value = row.get(column_name)
        if value is not None and not pd.isna(value) and str(value).strip():
            return str(value).strip()
    return f"ROW-{row.name}"


def _build_hengyi_summary_sheet(result_df: pd.DataFrame) -> pd.DataFrame:
    working = result_df.copy()
    working["订单号"] = working.apply(_build_hengyi_order_key, axis=1)
    if "托盘数" in working.columns:
        working["托盘数"] = working["托盘数"].fillna(0)

    summary_df = (
        working.groupby(["异常类型", "订单号"], as_index=False)
        .agg(
            {
                "工厂": _first_non_empty_value,
                "会员名称": _first_non_empty_value,
                "送达方": _first_non_empty_value,
                "托盘数": "sum",
                "出库数量": _first_non_empty_value,
                "出库数量差异": _first_non_empty_value,
            }
        )
    )
    summary_df["公司"] = summary_df["会员名称"].combine_first(summary_df["送达方"])
    summary_df = summary_df[
        ["异常类型", "订单号", "工厂", "公司", "托盘数", "出库数量", "出库数量差异"]
    ]
    anomaly_order = {"工厂缺单": 0, "久鼎缺单": 1, "数量差异": 2}
    summary_df["__sort_order"] = summary_df["异常类型"].map(anomaly_order).fillna(len(anomaly_order))
    summary_df = summary_df.sort_values(["__sort_order", "订单号"], kind="stable").drop(columns="__sort_order")
    summary_df = summary_df.reset_index(drop=True)
    return summary_df


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    export_df = df.copy().astype(object)
    export_df = export_df.where(pd.notna(export_df), None)
    export_df.to_excel(writer, sheet_name=sheet_name, index=False)


def _write_hengyi_detail_sheet(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    detail_df = df.copy()
    detail_df = detail_df.reindex(columns=HENGYI_DETAIL_COLUMNS)
    detail_df = detail_df.astype(object).where(pd.notna(detail_df), None)
    detail_df.to_excel(writer, sheet_name="异常详情", index=False, startrow=1)

    worksheet = writer.book["异常详情"]
    worksheet.freeze_panes = "A3"

    column_lookup = {column_name: index + 1 for index, column_name in enumerate(HENGYI_DETAIL_COLUMNS)}
    worksheet.cell(row=1, column=column_lookup["异常类型"], value="异常")
    worksheet.cell(row=1, column=column_lookup["出库数量差异"], value="差异")
    for group_name, columns in HENGYI_HEADER_GROUPS.items():
        start_column = column_lookup[columns[0]]
        end_column = column_lookup[columns[-1]]
        worksheet.cell(row=1, column=start_column, value=group_name)
        worksheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)


def _save_hengyi_result_workbook(file_path: str, result_df: pd.DataFrame) -> None:
    summary_df = _build_hengyi_summary_sheet(result_df)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        _write_sheet(writer, "异常汇总", summary_df)
        _write_hengyi_detail_sheet(writer, result_df)


def _build_hengyi_result_bytes(result_df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    summary_df = _build_hengyi_summary_sheet(result_df)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_sheet(writer, "异常汇总", summary_df)
        _write_hengyi_detail_sheet(writer, result_df)

    return buffer.getvalue()


def _resolve_xinfengming_anomaly_type(row: pd.Series) -> str:
    factory_qty = int(row.get("客户出库数") or 0)
    jiuding_qty = int(row.get("久鼎出库数") or 0)
    if factory_qty > 0 and jiuding_qty > 0:
        return "数量差异"
    if factory_qty > 0:
        return "久鼎缺单"
    return "工厂缺单"


def _build_xinfengming_detail_sheet(result_df: pd.DataFrame, artifacts: Dict[str, Any]) -> pd.DataFrame:
    factory_records = pd.DataFrame(artifacts.get("factory_records", []))
    jiuding_records = pd.DataFrame(artifacts.get("jiuding_records", []))

    detail_rows: list[dict[str, Any]] = []
    for _, summary_row in result_df.iterrows():
        order_no = summary_row.get("单号")
        anomaly_type = _resolve_xinfengming_anomaly_type(summary_row)

        factory_matches = (
            factory_records[factory_records["单号"] == order_no].to_dict(orient="records")
            if not factory_records.empty and "单号" in factory_records.columns
            else []
        )
        jiuding_matches = (
            jiuding_records[jiuding_records["单号"] == order_no].to_dict(orient="records")
            if not jiuding_records.empty and "单号" in jiuding_records.columns
            else []
        )
        max_rows = max(len(factory_matches), len(jiuding_matches), 1)

        for index in range(max_rows):
            factory_detail = factory_matches[index] if index < len(factory_matches) else {}
            jiuding_detail = jiuding_matches[index] if index < len(jiuding_matches) else {}
            detail_rows.append(
                {
                    "异常类型": anomaly_type,
                    "交货单号": factory_detail.get("交货单号"),
                    "交货创建日期": factory_detail.get("交货创建日期"),
                    "销售组织描述": factory_detail.get("销售组织描述"),
                    "客户名称": factory_detail.get("客户名称"),
                    "业务员": factory_detail.get("业务员"),
                    "车牌号": factory_detail.get("车牌号"),
                    "物料组描述": factory_detail.get("物料组描述"),
                    "件数": factory_detail.get("件数"),
                    "交货单类型": factory_detail.get("交货单类型"),
                    "包装批号": factory_detail.get("包装批号"),
                    "出库单号": jiuding_detail.get("出库单号"),
                    "订单日期": jiuding_detail.get("订单日期"),
                    "客户名称2": jiuding_detail.get("客户名称"),
                    "会员名称": jiuding_detail.get("会员名称"),
                    "产品类型": jiuding_detail.get("产品类型"),
                    "实际出库数量": jiuding_detail.get("实际出库数量"),
                    "子公司名称": jiuding_detail.get("子公司名称"),
                    "订单状态": jiuding_detail.get("订单状态"),
                    "送货方式": jiuding_detail.get("送货方式"),
                    "差异数量": summary_row.get("待处理数量"),
                }
            )

    return pd.DataFrame(detail_rows, columns=XINFENGMING_DETAIL_COLUMNS)


def _write_xinfengming_detail_sheet(writer: pd.ExcelWriter, detail_df: pd.DataFrame) -> None:
    export_df = detail_df.copy().astype(object)
    export_df = export_df.where(pd.notna(export_df), None)
    export_df.to_excel(writer, sheet_name="异常详情", index=False, startrow=1)

    worksheet = writer.book["异常详情"]
    worksheet.freeze_panes = "A3"

    column_lookup = {column_name: index + 1 for index, column_name in enumerate(XINFENGMING_DETAIL_COLUMNS)}
    worksheet.cell(row=1, column=column_lookup["异常类型"], value="异常")
    for group_name, columns in XINFENGMING_HEADER_GROUPS.items():
        start_column = column_lookup[columns[0]]
        end_column = column_lookup[columns[-1]]
        worksheet.cell(row=1, column=start_column, value=group_name)
        if start_column != end_column:
            worksheet.merge_cells(start_row=1, start_column=start_column, end_row=1, end_column=end_column)


def _build_xinfengming_result_bytes(result_df: pd.DataFrame, artifacts: Dict[str, Any]) -> bytes:
    buffer = io.BytesIO()
    summary_df = result_df.reindex(columns=XINFENGMING_SUMMARY_COLUMNS)
    detail_df = _build_xinfengming_detail_sheet(result_df, artifacts)

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_sheet(writer, "异常汇总", summary_df)
        _write_xinfengming_detail_sheet(writer, detail_df)
    return buffer.getvalue()


def _is_hengyi_result(result_df: pd.DataFrame, artifacts: Dict[str, Any]) -> bool:
    if artifacts.get("factory_type") == "hengyi":
        return True
    return "异常类型" in result_df.columns


def _save_result(
    *,
    result_df: pd.DataFrame,
    artifacts: Dict[str, Any],
) -> Dict[str, Any]:
    filename = _build_result_filename(result_df, factory_type=artifacts.get("factory_type"))
    if _is_hengyi_result(result_df, artifacts):
        download_bytes = _build_hengyi_result_bytes(result_df)
    else:
        download_bytes = _build_xinfengming_result_bytes(result_df, artifacts)

    download_token = base64.b64encode(download_bytes).decode("ascii")

    return {
        "data": result_df.to_dict(orient="records"),
        "filename": filename,
        "total_count": len(result_df),
        "download_token": download_token,
        "artifacts": artifacts,
    }


def _run_hengyi_comparison_sync(
    *,
    task_id: str,
    factory_files: List[Dict[str, Any]],
    jiuding_files: List[Dict[str, Any]],
) -> Dict[str, Any]:
    _update_task(task_id, status="parsing", progress=15, message="正在解析恒逸工厂数据...")

    parsed_factory_files = []
    selected_short_names: set[str] = set()
    for file_item in factory_files:
        source_df = _read_excel_with_fallback(file_item["content"])
        parsed_df = parse_hengyi_factory_data(source_df, source_filename=file_item["filename"])
        selected_short_names.update(
            short_name
            for short_name in parsed_df["工厂简称"].dropna().tolist()
            if str(short_name).strip()
        )
        LOGGER.info(
            "hengyi_factory filename=%s columns=%s selected_factories=%s",
            file_item["filename"],
            ["过账日期", "送达方", "交货单", "车牌号", "托盘数", "工厂"],
            sorted(selected_short_names),
        )
        parsed_factory_files.append({"filename": file_item["filename"], "parsed_df": parsed_df})

    _update_task(task_id, progress=45, message="正在筛选久鼎数据...")
    parsed_jiuding_files = []
    for file_item in jiuding_files:
        source_df = _read_excel_with_fallback(file_item["content"])
        parsed_df = parse_hengyi_jiuding_data(
            source_df,
            selected_factory_short_names=selected_short_names or None,
        )
        LOGGER.info(
            "hengyi_jiuding filename=%s columns=%s selected_factories=%s",
            file_item["filename"],
            ["订单日期", "出库单号", "会员名称", "客户名称", "产品类型", "实际出库数量"],
            sorted(selected_short_names),
        )
        parsed_jiuding_files.append({"filename": file_item["filename"], "parsed_df": parsed_df})

    factory_df = pd.concat(
        [item["parsed_df"] for item in parsed_factory_files],
        ignore_index=True,
    ) if parsed_factory_files else pd.DataFrame()
    jiuding_df = pd.concat(
        [item["parsed_df"] for item in parsed_jiuding_files],
        ignore_index=True,
    ) if parsed_jiuding_files else pd.DataFrame()

    _update_task(task_id, status="comparing", progress=75, message="正在汇总恒逸异常订单...")
    result_df = compare_hengyi_data(factory_df, jiuding_df)

    _update_task(task_id, progress=90, message="正在生成结果文件...")
    return _save_result(
        result_df=result_df,
        artifacts={
            "factory_type": "hengyi",
            "factory_files": [
                {
                    "filename": item["filename"],
                    "preview": "",
                    "plan": {
                        "mode": "fixed_columns",
                        "fields": ["过账日期", "送达方", "交货单", "车牌号", "托盘数", "工厂", "交货数量", "业务员", "物料组"],
                    },
                }
                for item in parsed_factory_files
            ],
            "jiuding_files": [
                {
                    "filename": item["filename"],
                    "preview": "",
                    "plan": {
                        "mode": "fixed_columns",
                        "fields": ["订单日期", "出库单号", "会员名称", "客户名称", "产品类型", "实际出库数量"],
                    },
                }
                for item in parsed_jiuding_files
            ],
        },
    )


def _run_xinfengming_comparison_sync(
    *,
    task_id: str,
    factory_files: List[Dict[str, Any]],
    jiuding_files: List[Dict[str, Any]],
    factory_type: str,
    llm_overrides: Dict[str, str],
) -> Dict[str, Any]:
    _update_task(task_id, status="parsing", progress=15, message="正在解析上传文件...")
    comparison_payload = compare_xinfengming_data(
        factory_files=factory_files,
        jiuding_files=jiuding_files,
        factory_type=factory_type,
    )

    _update_task(task_id, status="comparing", progress=75, message="正在汇总订单并核对差异...")
    result_df = comparison_payload["result_df"]

    _update_task(task_id, progress=90, message="正在生成结果文件...")
    return _save_result(
        result_df=result_df,
        artifacts=comparison_payload["artifacts"],
    )


def _run_comparison_sync(
    *,
    task_id: str,
    factory_files: List[Dict[str, Any]],
    jiuding_files: List[Dict[str, Any]],
    factory_type: str,
    llm_overrides: Dict[str, str],
) -> Dict[str, Any]:
    if factory_type == "hengyi":
        return _run_hengyi_comparison_sync(
            task_id=task_id,
            factory_files=factory_files,
            jiuding_files=jiuding_files,
        )

    return _run_xinfengming_comparison_sync(
        task_id=task_id,
        factory_files=factory_files,
        jiuding_files=jiuding_files,
        factory_type=factory_type,
        llm_overrides=llm_overrides,
    )


async def process_comparison(
    *,
    task_id: str,
    factory_files: List[Dict[str, Any]],
    jiuding_files: List[Dict[str, Any]],
    factory_type: str,
    llm_overrides: Dict[str, str],
) -> Dict[str, Any]:
    return await asyncio.to_thread(
        _run_comparison_sync,
        task_id=task_id,
        factory_files=factory_files,
        jiuding_files=jiuding_files,
        factory_type=factory_type,
        llm_overrides=llm_overrides,
    )


async def _run_comparison_task(
    *,
    task_id: str,
    factory_files: List[Dict[str, Any]],
    jiuding_files: List[Dict[str, Any]],
    factory_type: str,
    llm_overrides: Dict[str, str],
) -> None:
    try:
        result = await process_comparison(
            task_id=task_id,
            factory_files=factory_files,
            jiuding_files=jiuding_files,
            factory_type=factory_type,
            llm_overrides=llm_overrides,
        )
        tasks[task_id]["result"] = _clean_data(result)
        _update_task(task_id, status="completed", progress=100, message="比对完成")
    except HengyiComparisonError as exc:
        _update_task(task_id, status="failed", message=f"处理失败: {exc}")
        LOGGER.exception("task=%s failed", task_id)
    except Exception as exc:
        _update_task(task_id, status="failed", message=f"处理失败: {exc}")
        LOGGER.exception("task=%s failed", task_id)
        print(traceback.format_exc())


@router.post("/compare")
async def create_comparison(
    factory_files: List[UploadFile] = File(...),
    jiuding_files: List[UploadFile] = File(...),
    factory_type: str = Form("hengyi"),
    llm_base_url: Optional[str] = Form(None),
    llm_api_key: Optional[str] = Form(None),
    llm_model: Optional[str] = Form(None),
    llm_transport: Optional[str] = Form(None),
):
    if not factory_files:
        raise HTTPException(status_code=400, detail="工厂侧文件不能为空")
    if not jiuding_files:
        raise HTTPException(status_code=400, detail="久鼎侧文件不能为空")

    for file_item in factory_files:
        _ensure_excel(file_item.filename, "工厂侧文件")
    for file_item in jiuding_files:
        _ensure_excel(file_item.filename, "久鼎侧文件")

    task_id = str(uuid.uuid4())
    tasks[task_id] = {
        "task_id": task_id,
        "status": "pending",
        "progress": 5,
        "message": "任务已创建，等待处理...",
        "result": None,
    }

    factory_payloads = [
        {"filename": file_item.filename, "content": await file_item.read()}
        for file_item in factory_files
    ]
    jiuding_payloads = [
        {"filename": file_item.filename, "content": await file_item.read()}
        for file_item in jiuding_files
    ]

    asyncio.create_task(
        _run_comparison_task(
            task_id=task_id,
            factory_files=factory_payloads,
            jiuding_files=jiuding_payloads,
            factory_type=factory_type,
            llm_overrides={
                "base_url": llm_base_url or "",
                "api_key": llm_api_key or "",
                "model": llm_model or "",
                "transport": llm_transport or "",
            },
        )
    )

    return {"task_id": task_id, "status": "pending"}


@router.get("/compare/{task_id}/status")
async def get_task_status(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")

    task = tasks[task_id]
    return TaskStatus(
        task_id=task["task_id"],
        status=task["status"],
        progress=task["progress"],
        message=task["message"],
    )


@router.get("/compare/{task_id}/result")
async def get_task_result(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    if tasks[task_id]["status"] != "completed":
        raise HTTPException(status_code=400, detail="任务尚未完成")
    return JSONResponse(content=_clean_data(tasks[task_id]["result"]))


@router.get("/compare/{task_id}/download")
async def download_result(task_id: str):
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    if tasks[task_id]["status"] != "completed":
        raise HTTPException(status_code=400, detail="任务尚未完成")

    filename = tasks[task_id]["result"]["filename"]
    download_token = tasks[task_id]["result"].get("download_token")
    if not download_token:
        raise HTTPException(status_code=404, detail="结果文件不存在")

    file_bytes = base64.b64decode(download_token)
    safe_ascii_filename = "download.xlsx"
    encoded_filename = quote(filename)
    content_disposition = (
        f"attachment; filename=\"{safe_ascii_filename}\"; "
        f"filename*=UTF-8''{encoded_filename}"
    )

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition},
    )


@router.get("/factory-groups")
async def get_factory_groups() -> Dict[str, Any]:
    with FACTORY_GROUPS_CONFIG_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)
